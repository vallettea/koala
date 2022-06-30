import openpyxl

def to_excel(spreadsheet, fname=None):
    '''chaz's thing'''

    if fname is None:
        raise Exception('No filename specified. Please provide one.')
    else:
        fname=fname.split('.')[0]

    # TODO:
    # sort sheets before creating
    # test ad-hoc worksheets
    # pivot tables?  eek.

    spreadsheet.prune_graph()
    #spreadsheet.clean_pointer()

    '''
    isolate sheets, cells, and formulae from graph into a dict:
        {thisSheet: [(thisCell, thisFormula), (thisCell_1, thisFormula_1),...]}
    do not include range names as keys
    '''

    theDict={}

    #print(spreadsheet.addr_to_name)
    #print(spreadsheet.addr_to_range)
    #print(spreadsheet.named_ranges)

    print('reading tree contents...')
    print('address, formula, value')

    for c in list(spreadsheet.cellmap.values()):
        print(c.address(), c.formula, c.value)
        '''
        actual name ranges (as opposed to named cells (single-cell ranges)) should be excluded
        from theDict
        '''
        if not any(c.address() in val for val in spreadsheet.addr_to_range.values()):
            #print(c.address(), c.formula, c. value)

            thisCell = None

            thisAddress = c.address().split('!')
            thisSheet = thisAddress[0]
            if len(thisAddress) > 1:
                thisCell = thisAddress[1]

            #thisCell=None
            # thisSheet = c.address().split('!')[0]
            # if len(c.address().)

            thisFormula = c.formula
            if thisFormula is not None and thisFormula.find('=') != 0: thisFormula = '=' + thisFormula

            thisValue = c.value

            if thisFormula is not None:
                thisValue = thisFormula

            print('collecting ' + thisSheet, thisCell, thisValue )
            if thisSheet not in theDict:
                theDict[thisSheet] = [(thisCell, thisValue)]
            else:
                theDict[thisSheet].append((thisCell, thisValue))

    '''
    clean up dict by removing range names from keys (keys are spreadsheet names)
    '''
    for i in spreadsheet.named_ranges:
        if i in theDict:
            print('      removing name range for special handling')
            theDict.pop(i)

    # print('------the dict------')
    # print(theDict)
    # print('--------------------')

    '''
    create the workbook with openpyxl
    '''
    wb = openpyxl.Workbook()

    '''
    create sheets from theDict
    '''
    for sheetName in theDict.keys():
        print('   creating sheet... ', sheetName)
        wb.create_sheet(sheetName)

    '''
    get rid of the sheet autocreated by openpyxl
    and don't consider range names (if any) as sheet named_ranges
    '''
    for sheet in wb.sheetnames:
        if sheet not in theDict.keys() or sheet in spreadsheet.named_ranges:
            rm_sheet =  wb[sheet]
            wb.remove(rm_sheet)

    '''
    add formuale to cells by sheets
    '''
    for sheetName, values in theDict.items():
        print('adding cell contents...')
        for cells in values:
            thisCell = cells[0]
            thisFormula = cells[1]
            print('   adding ' + sheetName + '!' + thisCell, thisFormula)
            wb[sheetName][thisCell] = thisFormula

    '''
    add named ranges
    '''
    print('adding ranges')
    for thisName, thisAddress in spreadsheet.named_ranges.items():
        thisAddress = absolute_addr(thisAddress)

        print('   adding range ', thisName, thisAddress)
        theRange = openpyxl.defined_name.DefinedName(thisName, attr_text=thisAddress)
        wb.defined_names.append(theRange)

    print('saving wb')
    try:
        wb.close()
        wb.save(fname + '.xlsx')
    except Exception as e:
        print('error saving wb: "' + str(e)+'"')

def absolute_addr(theAddress):
    #make the address absolute
    theSheet = theAddress.split('!')[0]
    theCells = theAddress.split('!')[1:]
    for cell in theCells:
        absCell =  ''
        for i in cell.split(':'):
            absCell = absCell + openpyxl.cell.absolute_coordinate(i.strip())
            if len(cell.split(':')) > 1: absCell = absCell + ':'
        if absCell[-1] == ':': absCell = absCell[:-1]
        absAddress = theSheet + '!' + absCell
    return absAddress
